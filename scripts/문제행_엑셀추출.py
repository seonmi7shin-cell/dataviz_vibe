# data/01_핀테크결제_dirty.csv 에서 문제가 있는 행(결측·완전중복·음수 거래금액)만 골라
# 어떤 문제인지 표시하고, 문제 셀에 색을 입혀 엑셀로 저장하는 스크립트.

import os
import pandas as pd
from openpyxl.styles import PatternFill, Font

os.chdir(os.path.dirname(os.path.abspath(__file__)) + "/..")

df = pd.read_csv(os.path.join("data", "01_핀테크결제_dirty.csv"), encoding="utf-8-sig")

# 세 가지 문제를 각각 행 단위로 판정
is_missing = df.isna().any(axis=1)
is_duplicate = df.duplicated(keep=False)  # 원본·복사본 모두 포함해서 표시
is_negative = df["거래금액"] < 0

has_problem = is_missing | is_duplicate | is_negative
problem_df = df[has_problem].copy()

# 각 행에 어떤 문제인지 사람이 읽을 수 있는 문구로 표시 (여러 개면 쉼표로 나열)
def reasons(row_idx):
    tags = []
    if is_missing[row_idx]:
        missing_cols = df.columns[df.loc[row_idx].isna()].tolist()
        tags.append("결측(" + ",".join(missing_cols) + ")")
    if is_duplicate[row_idx]:
        tags.append("완전중복")
    if is_negative[row_idx]:
        tags.append("음수거래금액")
    return " / ".join(tags)

problem_df.insert(0, "문제유형", [reasons(i) for i in problem_df.index])

print(f"전체 행 수: {len(df)}")
print(f"결측 포함 행: {is_missing.sum()}건")
print(f"완전 중복 행: {is_duplicate.sum()}건")
print(f"음수 거래금액 행: {is_negative.sum()}건")
print(f"문제 있는 행(중복 제외 합집합): {len(problem_df)}건")

output_dir = "output"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "문제행_점검.xlsx")
problem_df.to_excel(output_path, index=False, sheet_name="문제행")

# 문제가 있는 셀만 색을 입혀서 어디가 문제인지 한눈에 보이게 함
from openpyxl import load_workbook

wb = load_workbook(output_path)
ws = wb["문제행"]

yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 결측 칸
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # 음수 거래금액
orange = PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid")  # 완전 중복 행

col_names = ["문제유형"] + list(df.columns)
col_index = {name: idx + 1 for idx, name in enumerate(col_names)}  # 엑셀은 1부터 시작

for excel_row, (orig_idx, row) in enumerate(problem_df.iterrows(), start=2):
    if is_duplicate[orig_idx]:
        for col in range(1, len(col_names) + 1):
            ws.cell(row=excel_row, column=col).fill = orange
    if is_missing[orig_idx]:
        for col_name in df.columns[df.loc[orig_idx].isna()]:
            ws.cell(row=excel_row, column=col_index[col_name]).fill = yellow
    if is_negative[orig_idx]:
        ws.cell(row=excel_row, column=col_index["거래금액"]).fill = red
        ws.cell(row=excel_row, column=col_index["거래금액"]).font = Font(bold=True)

wb.save(output_path)
print(f"저장 경로: {output_path}")
