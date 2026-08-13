import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

WEEKDAY_KR = ['월', '화', '수', '목', '금', '토', '일']

# (기존 과정명, 2027 과정명, 구분, 교육일수, [(시작일, 종료일), ...])
courses = [
    (
        "Claude Code를 활용한 데이터 분석 자동화 & AI 대시보드",
        "Claude Code를 활용한 데이터 분석 자동화 & AI 대시보드",
        "여의도", 2,
        [
            ("2027-01-13", "2027-01-14"),
            ("2027-03-04", "2027-03-05"),
            ("2027-04-15", "2027-04-16"),
            ("2027-06-16", "2027-06-17"),
            ("2027-07-22", "2027-07-23"),
            ("2027-09-08", "2027-09-09"),
            ("2027-10-12", "2027-10-13"),
            ("2027-11-11", "2027-11-12"),
        ],
    ),
    (
        "바이브코딩으로 구축하는 사내 AI 챗봇·지식 에이전트(with Claude Code)",
        "바이브코딩으로 구축하는 사내 AI 챗봇·지식 에이전트(with Claude Code)",
        "여의도", 2,
        [
            ("2027-01-07", "2027-01-08"),
            ("2027-03-24", "2027-03-25"),
            ("2027-05-27", "2027-05-28"),
            ("2027-07-15", "2027-07-16"),
            ("2027-10-07", "2027-10-08"),
            ("2027-11-24", "2027-11-25"),
        ],
    ),
]


def weekday_kr(date_str):
    d = datetime.date.fromisoformat(date_str)
    return WEEKDAY_KR[d.weekday()]


rows = []
for existing_name, name_2027, region, days, sessions in courses:
    for start, end in sessions:
        s_dt = datetime.date.fromisoformat(start)
        e_dt = datetime.date.fromisoformat(end)
        rows.append([
            existing_name,
            name_2027,
            region,
            days,
            start,
            weekday_kr(start),
            end,
            weekday_kr(end),
            f"{s_dt.month}월 {s_dt.day}~{e_dt.day}일 ({weekday_kr(start)}~{weekday_kr(end)})",
        ])

# 시작일 기준으로 정렬. 두 과정이 섞여서 시간 순으로 이어져야 전체 일정표로 보기 편함
rows.sort(key=lambda r: r[4])

wb = Workbook()
ws = wb.active
ws.title = "2027 교육일정"

headers = ["기존 과정명", "2027 과정명", "구분", "교육일수", "시작일", "시작 요일", "종료일", "종료 요일", "일정 표기"]
ws.append(headers)

for row in rows:
    ws.append(row)

# 헤더 스타일: 굵게 + 배경색을 넣어 표의 첫 행임을 한눈에 알아보게 함
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# 화요일·수요일에 걸리는 요일 셀만 강조. 여의도 교육이 보통 목~금인데 예외로 화~수에 걸리는
# 회차를 담당자가 한눈에 구분할 수 있게 시작/종료 요일 칸에만 색을 입힘
highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
highlight_font = Font(bold=True, color="9C5700")

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # 시작 요일(F열) 또는 종료 요일(H열) 중 하나라도 월/화면 두 요일 셀을 함께 강조.
    # 한쪽만 강조하면 같은 회차인데 왜 한 칸만 표시가 다른지 헷갈릴 수 있어서 짝으로 처리함
    start_day_cell = row[5]
    end_day_cell = row[7]
    has_mon_tue = start_day_cell.value in ("월", "화") or end_day_cell.value in ("월", "화")

    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="center")

    if has_mon_tue:
        for cell in (start_day_cell, end_day_cell):
            cell.fill = highlight_fill
            cell.font = highlight_font

# 열 너비: 과정명처럼 긴 텍스트가 잘리지 않도록 넉넉히 잡음
widths = [40, 40, 8, 8, 12, 10, 12, 10, 30]
for idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(idx)].width = w

ws.freeze_panes = "A2"

out_path = "KMA_2027_교육일정_최종.xlsx"
wb.save(out_path)
print(f"저장 완료: {out_path}")
print(f"총 {len(rows)}건의 교육 일정 정리")
